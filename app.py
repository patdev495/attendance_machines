import argparse
import os
import uvicorn
from fastapi import FastAPI, Depends, Query, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from starlette.background import BackgroundTask
import tempfile
from sqlalchemy.orm import Session
from sqlalchemy import func, desc, text, Date, case, or_
from typing import List, Optional
from datetime import date, datetime, time, timedelta
from db import get_db, AttendanceLog, EmployeeMetadata
from sync import sync_all_machines, get_machine_list, sync_status, sync_employees_from_excel, delete_user_from_all_machines, get_users_from_machine, delete_user_from_machine

app = FastAPI(title="Time Attendance System")

# CORS middleware to allow frontend to access backend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Shift Rule Helpers ────────────────────────────────────────────
XUONG1_DEPT_KEYWORD = "Xưởng 1"

def get_shift_rules(department, shift):
    """Return shift rules dict based on department + shift type."""
    is_xuong1 = bool(department and XUONG1_DEPT_KEYWORD in str(department))
    if is_xuong1:
        if shift == 'D':  # Night: 20:00 → 08:00 next day
            return dict(official_start=time(20,0), official_end=time(8,0), end_next_day=True,
                        max_hours=12.0, standard_hours=12.0, deduct_break=False, has_overtime=False)
        else:              # Day:   08:00 → 20:00
            return dict(official_start=time(8,0),  official_end=time(20,0), end_next_day=False,
                        max_hours=12.0, standard_hours=12.0, deduct_break=False, has_overtime=False)
    else:
        if shift == 'D':  # Night: 20:00 → 05:00 next day
            return dict(official_start=time(20,0), official_end=time(5,0),  end_next_day=True,
                        max_hours=None,  standard_hours=8.0,  deduct_break=True,  has_overtime=True)
        else:              # Day:   08:00 → 17:00
            return dict(official_start=time(8,0),  official_end=time(17,0), end_next_day=False,
                        max_hours=None,  standard_hours=8.0,  deduct_break=True,  has_overtime=True)

def compute_day_stats(first, last, w_date, department, shift):
    """Return (work_hours, hours_standard, hours_ot, minutes_late, minutes_early)."""
    rules = get_shift_rules(department, shift)
    official_start_dt = datetime.combine(w_date, rules['official_start'])
    official_end_dt   = (datetime.combine(w_date + timedelta(days=1), rules['official_end'])
                         if rules['end_next_day']
                         else datetime.combine(w_date, rules['official_end']))

    minutes_late  = max(0, int((first - official_start_dt).total_seconds() / 60))
    minutes_early = max(0, int((official_end_dt - last).total_seconds() / 60))

    effective_in  = max(first, official_start_dt)
    effective_out = min(last, official_end_dt) if not rules['has_overtime'] else last
    total_secs    = max(0.0, (effective_out - effective_in).total_seconds())

    if rules['deduct_break']:
        work_hours = round((total_secs - 3600) / 3600, 2) if total_secs > 3600 else round(total_secs / 3600, 2)
    else:
        work_hours = round(total_secs / 3600, 2)

    if rules['max_hours'] is not None:
        work_hours = min(work_hours, rules['max_hours'])

    if rules['has_overtime'] and work_hours >= rules['standard_hours']:
        hours_standard = rules['standard_hours']
        hours_ot       = round(work_hours - rules['standard_hours'], 2)
    else:
        hours_standard = work_hours
        hours_ot       = 0.0

    return work_hours, hours_standard, hours_ot, minutes_late, minutes_early
# ──────────────────────────────────────────────────────────────────

@app.get("/api/machines")
def get_machines():
    return get_machine_list()

@app.get("/api/attendance")
def get_attendance(
    employee_id: Optional[str] = Query(None),
    machine_ip: Optional[str] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1),
    db: Session = Depends(get_db)
):
    query = db.query(AttendanceLog).outerjoin(EmployeeMetadata, AttendanceLog.employee_id == EmployeeMetadata.employee_id)
    
    if employee_id:
        query = query.filter(AttendanceLog.employee_id == employee_id)
    if machine_ip:
        query = query.filter(AttendanceLog.machine_ip == machine_ip)
    if start_date:
        query = query.filter(AttendanceLog.attendance_date >= start_date)
    if end_date:
        query = query.filter(AttendanceLog.attendance_date <= end_date)
    if status:
        query = query.filter(EmployeeMetadata.status == status)
        
    total_count = query.count()
    total_pages = (total_count + size - 1) // size
    
    results = query.order_by(desc(AttendanceLog.attendance_time)) \
                   .offset((page - 1) * size) \
                   .limit(size) \
                   .all()
    
    return {
        "items": results,
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "size": size
    }

@app.get("/api/attendance/summary")
def get_attendance_summary(
    employee_id: Optional[str] = Query(None),
    machine_ip: Optional[str] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    min_hours: Optional[float] = Query(None),
    max_hours: Optional[float] = Query(None),
    only_missing: bool = Query(False),
    shift: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1),
    db: Session = Depends(get_db)
):
    # Shift-specific Date Logic:
    # Night Shift (D): We offset by 10 hours so an 08:00 AM tap is counted as 'Yesterday'.
    # Day Shift (N): We offset by 4 hours for a normal morning boundary.
    # Step 1: Base subquery with join and date calculation
    # We use -10 for Night (D) and -4 for Day (N/All)
    base_calc_sub = db.query(
        AttendanceLog.id,
        AttendanceLog.attendance_time,
        AttendanceLog.employee_id,
        AttendanceLog.machine_ip,
        EmployeeMetadata.shift,
        EmployeeMetadata.status,
        EmployeeMetadata.department,
        case(
            (EmployeeMetadata.shift == text("'D'"), func.cast(func.dateadd(text("hour"), text("-10"), AttendanceLog.attendance_time), Date)),
            else_=func.cast(func.dateadd(text("hour"), text("-4"), AttendanceLog.attendance_time), Date)
        ).label("work_date")
    ).outerjoin(EmployeeMetadata, AttendanceLog.employee_id == EmployeeMetadata.employee_id).subquery()
    
    # Step 2: Aggregated query on top of the subquery
    query = db.query(
        base_calc_sub.c.employee_id,
        base_calc_sub.c.work_date,
        func.min(base_calc_sub.c.attendance_time).label("first_tap"),
        func.max(base_calc_sub.c.attendance_time).label("last_tap"),
        func.count(base_calc_sub.c.id).label("tap_count"),
        base_calc_sub.c.shift,
        base_calc_sub.c.status,
        base_calc_sub.c.department
    )
    
    if employee_id:
        query = query.filter(base_calc_sub.c.employee_id == employee_id)
    if machine_ip:
        query = query.filter(base_calc_sub.c.machine_ip == machine_ip)
    if start_date:
        query = query.filter(base_calc_sub.c.work_date >= start_date)
    if end_date:
        query = query.filter(base_calc_sub.c.work_date <= end_date)
    if shift:
        if shift == "NA":
            query = query.filter(or_(base_calc_sub.c.shift == None, base_calc_sub.c.shift == ""))
        else:
            query = query.filter(base_calc_sub.c.shift == shift)
    if status:
        query = query.filter(base_calc_sub.c.status == status)
        
    query = query.group_by(
        base_calc_sub.c.employee_id,
        base_calc_sub.c.work_date,
        base_calc_sub.c.shift,
        base_calc_sub.c.status,
        base_calc_sub.c.department
    )
    
    # Advanced Filtering using HAVING
    if only_missing:
        query = query.having(func.count(base_calc_sub.c.id) == 1)
    else:
        # If filtering by hours, the record must have at least 2 taps (In and Out)
        if min_hours is not None:
            # We calculate seconds difference and divide by 3600 for precision
            # Note: Cast to float for division
            query = query.having(
                (func.datediff(text("second"), func.min(base_calc_sub.c.attendance_time), func.max(base_calc_sub.c.attendance_time)) / 3600.0) >= min_hours
            )
        if max_hours is not None:
            query = query.having(
                (func.datediff(text("second"), func.min(base_calc_sub.c.attendance_time), func.max(base_calc_sub.c.attendance_time)) / 3600.0) <= max_hours
            )
    
    # Use a subquery to correctly count grouped rows
    total_count = db.query(func.count()).select_from(query.subquery()).scalar()
    total_pages = (total_count + size - 1) // size
    
    # Order by work_date desc, then employee
    results = query.order_by(desc(base_calc_sub.c.work_date), base_calc_sub.c.employee_id) \
                   .offset((page - 1) * size) \
                   .limit(size) \
                   .all()
    
    summary_items = []
    for row in results:
        first      = row.first_tap
        last       = row.last_tap
        count      = row.tap_count
        w_date     = row.work_date
        row_shift  = row.shift
        department = row.department

        work_hours       = 0.0
        minutes_late     = None
        minutes_early_lv = None
        note = ""

        if count > 1 and first != last:
            work_hours, _, _, minutes_late, minutes_early_lv = compute_day_stats(
                first, last, w_date, department, row_shift
            )
        else:
            note = "Missing Check-in/out (Only 1 tap)"

        summary_items.append({
            "employee_id": row.employee_id,
            "attendance_date": w_date,
            "first_tap": first,
            "last_tap": last,
            "tap_count": count,
            "work_hours": work_hours,
            "shift": row_shift or "N/A",
            "status": row.status or "Active",
            "note": note,
            "minutes_late": minutes_late,
            "minutes_early_leave": minutes_early_lv,
        })
        
    return {
        "items": summary_items,
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "size": size
    }

@app.post("/api/sync")
def trigger_sync(background_tasks: BackgroundTasks):
    if sync_status["is_running"]:
        return {"message": "Sync is already in progress.", "is_running": True}
    
    background_tasks.add_task(sync_all_machines)
    return {"message": "Sync started in background.", "is_running": True}

@app.get("/api/devices/capacity")
def get_devices_capacity():
    from sync import get_devices_capacity_info
    return get_devices_capacity_info()

@app.get("/api/sync-status")
def get_sync_status():
    from sync import sync_status
    return sync_status

@app.get("/api/employees/delete-status")
def get_delete_status():
    from sync import delete_status
    return delete_status

@app.post("/api/employees/sync")
def sync_employees():
    count, msg = sync_employees_from_excel()
    if msg != "Success":
        raise HTTPException(status_code=500, detail=msg)
    return {"message": f"Successfully synced {count} employees from Excel."}

@app.delete("/api/employees/{employee_id}/machine-data")
def delete_employee_from_machines(employee_id: str, background_tasks: BackgroundTasks):
    from sync import delete_user_from_all_machines, delete_status
    if delete_status["is_running"]:
        return {"message": "A deletion is already in progress.", "is_running": True}
        
    background_tasks.add_task(delete_user_from_all_machines, employee_id)
    return {"message": f"Deletion of employee {employee_id} started.", "is_running": True}

@app.get("/api/devices/{ip}/employees")
def get_device_employees(
    ip: str, 
    page: int = Query(1, ge=1), 
    size: int = Query(50, ge=1),
    db: Session = Depends(get_db)
):
    users, msg = get_users_from_machine(ip)
    if msg != "Success":
        raise HTTPException(status_code=500, detail=msg)
    
    # Merge with database metadata to get names/status
    emp_ids = [u["user_id"] for u in users]
    meta = {m.employee_id: m for m in db.query(EmployeeMetadata).filter(EmployeeMetadata.employee_id.in_(emp_ids)).all()}
    
    for u in users:
        m = meta.get(u["user_id"])
        if m:
            u["db_name"] = m.emp_name
            u["status"] = m.status
            u["department"] = m.department
        else:
            u["db_name"] = None
            u["status"] = "Unknown"
            u["department"] = None

    total_count = len(users)
    total_pages = (total_count + size - 1) // size
    start = (page - 1) * size
    end = start + size
    
    return {
        "items": users[start:end],
        "total_count": total_count,
        "total_pages": total_pages,
        "current_page": page,
        "size": size
    }

@app.delete("/api/devices/{ip}/employees/{employee_id}")
def delete_device_employee(ip: str, employee_id: str):
    result = delete_user_from_machine(ip, employee_id)
    if result == "Success":
        return {"message": f"Successfully deleted employee {employee_id} from machine {ip}"}
    elif result == "Not in device":
        raise HTTPException(status_code=404, detail="Employee not found on this device")
    else:
        raise HTTPException(status_code=500, detail=result)

@app.get("/api/export-attendance")
def export_attendance(
    start_date: date = Query(...),
    end_date: date = Query(...),
    view_mode: str = Query(..., description="'time' or 'hours'"),
    db: Session = Depends(get_db)
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    # 1. Base query similar to summary endpoint
    base_calc_sub = db.query(
        AttendanceLog.id,
        AttendanceLog.attendance_time,
        AttendanceLog.employee_id,
        EmployeeMetadata.shift,
        case(
            (EmployeeMetadata.shift == text("'D'"), func.cast(func.dateadd(text("hour"), text("-10"), AttendanceLog.attendance_time), Date)),
            else_=func.cast(func.dateadd(text("hour"), text("-4"), AttendanceLog.attendance_time), Date)
        ).label("work_date")
    ).outerjoin(EmployeeMetadata, AttendanceLog.employee_id == EmployeeMetadata.employee_id).subquery()
    
    query = db.query(
        base_calc_sub.c.employee_id,
        base_calc_sub.c.work_date,
        func.min(base_calc_sub.c.attendance_time).label("first_tap"),
        func.max(base_calc_sub.c.attendance_time).label("last_tap"),
        func.count(base_calc_sub.c.id).label("tap_count"),
        base_calc_sub.c.shift
    )
    query = query.filter(base_calc_sub.c.work_date >= start_date)
    query = query.filter(base_calc_sub.c.work_date <= end_date)
    query = query.group_by(
        base_calc_sub.c.employee_id, 
        base_calc_sub.c.work_date, 
        base_calc_sub.c.shift
    )
    
    results = query.all()
    
    # Process results into a dictionary: data[emp_id][date] = stats
    data = {}
    for row in results:
        emp_id = row.employee_id
        w_date = row.work_date
        
        if emp_id not in data:
            data[emp_id] = {
                "general_shift": row.shift,
                "days": {}
            }
            
        data[emp_id]["days"][w_date] = {
            "first_tap": row.first_tap,
            "last_tap": row.last_tap,
            "count": row.tap_count,
            "shift": row.shift
        }
        
    # Generate Date Columns
    dates_list = []
    curr = start_date
    while curr <= end_date:
        dates_list.append(curr)
        curr += timedelta(days=1)
        
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Export"
    
    # Header Row
    headers = ["Mã nhân viên", "Tên nhân viên", "Phòng ban", "Nhóm", "Ngày vào làm", "Ca làm việc", "Ghi chú"]
    for d in dates_list:
        headers.append(f"{d.day}/{d.month}")
        
    ws.append(headers)
    header_row = ws[1]
    
    border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
        cell.fill = header_fill
        
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    
    for col_idx in range(8, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
        
    # Sort data by employee_id for better output
    sorted_emp_ids = sorted(data.keys())
    
    emp_meta = {m.employee_id: m for m in db.query(EmployeeMetadata).filter(EmployeeMetadata.employee_id.in_(sorted_emp_ids)).all()}
    
    # Data Rows
    current_row = 2
    # time → 4 rows (In/Out + Đi muộn/Về sớm)
    # hours → 4 rows (Công/TăngCa + Đi muộn/Về sớm)
    # both  → 6 rows
    if view_mode == "both":
        num_rows = 6
    else:
        num_rows = 4
    
    for emp_id in sorted_emp_ids:
        emp_data_info = data[emp_id]
        emp_data = emp_data_info["days"]
        shift_val = emp_data_info["general_shift"]
        emp_m = emp_meta.get(emp_id)
        emp_name = emp_m.emp_name if emp_m and emp_m.emp_name else emp_id
        emp_dept = emp_m.department if emp_m and emp_m.department else "-"
        emp_group = emp_m.group if emp_m and emp_m.group else "-"
        emp_start = emp_m.start_date.strftime("%d/%m/%Y") if emp_m and emp_m.start_date else "-"
        
        if shift_val == 'D':
            shift_text = "D"
        elif shift_val == 'N':
            shift_text = "N"
        else:
            shift_text = shift_val or "-"
        
        # Write values to all rows for this employee so Excel Filtering works perfectly without merged cells
        for r_idx in range(current_row, current_row+num_rows):
            is_first = (r_idx == current_row)
            
            c1 = ws.cell(row=r_idx, column=1, value=emp_id)
            c2 = ws.cell(row=r_idx, column=2, value=emp_name)
            c3 = ws.cell(row=r_idx, column=3, value=emp_dept)
            c4 = ws.cell(row=r_idx, column=4, value=emp_group)
            c5 = ws.cell(row=r_idx, column=5, value=emp_start)
            c6 = ws.cell(row=r_idx, column=6, value=shift_text)
            
            # Indicator column
            if view_mode == "time":
                all_indicators = ["Giờ In", "Giờ Out", "Đi muộn (phút)", "Về sớm (phút)"]
            elif view_mode == "hours":
                all_indicators = ["Giờ Công", "Tăng Ca", "Đi muộn (phút)", "Về sớm (phút)"]
            else:
                all_indicators = ["Giờ In", "Giờ Out", "Giờ Công", "Tăng Ca", "Đi muộn (phút)", "Về sớm (phút)"]
            indicator = all_indicators[r_idx - current_row]
            c7 = ws.cell(row=r_idx, column=7, value=indicator)

            # Apply font white to duplicate employee info to look cleaner, but keep the data for filtering
            if not is_first:
                for c in [c1, c2, c3, c4, c5, c6]:
                    c.font = Font(color="FFFFFF")
            
            # borders and alignment
            top_b = 'thin' if is_first else None
            bottom_b = 'thin' if (r_idx == current_row + num_rows - 1) else None
            
            for c_idx in range(1, 8):
                c = ws.cell(row=r_idx, column=c_idx)
                c.alignment = Alignment(horizontal="center", vertical="center")
                # For G (indicator), keep full borders, others keep outer borders
                if c_idx == 7:
                    c.border = border_style
                else:
                    c.border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                                      top=Side(style=top_b) if top_b else None, 
                                      bottom=Side(style=bottom_b) if bottom_b else None)

        # Fill date columns
        for date_idx, d in enumerate(dates_list):
            col = date_idx + 8
            cells = [ws.cell(row=current_row+i, column=col) for i in range(num_rows)]
            
            for c in cells:
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border_style
            
            if d not in emp_data:
                for c in cells: c.value = "-"
            else:
                day_stat = emp_data[d]
                c_count  = day_stat["count"]
                emp_dept = emp_m.department if emp_m else None

                first_val      = "-"
                last_val       = "-"
                hours_std_val  = "-"
                hours_ot_val   = "-"
                late_val       = "-"
                early_val      = "-"

                if c_count >= 1 and day_stat["first_tap"]:
                    first_val = day_stat["first_tap"].strftime("%H:%M")

                if c_count >= 2 and day_stat["first_tap"] != day_stat["last_tap"]:
                    last_val = day_stat["last_tap"].strftime("%H:%M")
                    _, hours_std, hours_ot, min_late, min_early = compute_day_stats(
                        day_stat["first_tap"], day_stat["last_tap"],
                        d, emp_dept, day_stat["shift"]
                    )
                    hours_std_val = hours_std
                    hours_ot_val  = hours_ot if hours_ot > 0 else "-"
                    late_val      = min_late
                    early_val     = min_early

                has2 = c_count >= 2
                if view_mode == "time":
                    cells[0].value = first_val
                    cells[1].value = last_val
                    cells[2].value = late_val  if has2 else "-"
                    cells[3].value = early_val if has2 else "-"
                elif view_mode == "hours":
                    cells[0].value = hours_std_val if has2 else "-"
                    cells[1].value = hours_ot_val  if has2 else "-"
                    cells[2].value = late_val      if has2 else "-"
                    cells[3].value = early_val     if has2 else "-"
                else:  # both
                    cells[0].value = first_val
                    cells[1].value = last_val
                    cells[2].value = hours_std_val if has2 else "-"
                    cells[3].value = hours_ot_val  if has2 else "-"
                    cells[4].value = late_val      if has2 else "-"
                    cells[5].value = early_val     if has2 else "-"
                            
        current_row += num_rows

    # ─── Sheet 2: Flat/Filterable "Late & Early Leave" table ─────────
    ws2 = wb.create_sheet(title="Thông tin chi tiết")

    flat_headers = [
        "Mã NV", "Tên nhân viên", "Phòng ban", "Nhóm", "Ca",
        "Ngày", "Giờ In", "Giờ Out", "Giờ Công", "Tăng Ca",
        "Đi muộn (phút)", "Về sớm (phút)"
    ]
    ws2.append(flat_headers)
    h2_row = ws2[1]
    for cell in h2_row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style
        cell.fill = header_fill

    # Column widths for sheet 2
    for col_letter, width in zip("ABCDEFGHIJKL", [12, 25, 20, 12, 6, 12, 10, 10, 10, 10, 16, 16]):
        ws2.column_dimensions[col_letter].width = width

    for emp_id in sorted_emp_ids:
        emp_data    = data[emp_id]["days"]
        emp_m       = emp_meta.get(emp_id)
        emp_name    = emp_m.emp_name   if emp_m and emp_m.emp_name   else emp_id
        emp_dept    = emp_m.department if emp_m and emp_m.department else "-"
        emp_group   = emp_m.group      if emp_m and emp_m.group      else "-"
        emp_dept_raw = emp_m.department if emp_m else None

        for d in dates_list:
            if d not in emp_data:
                continue
            day_stat = emp_data[d]
            c_count  = day_stat["count"]
            row_shift = day_stat["shift"] or "N"
            shift_text = "D" if row_shift == "D" else "N"

            first_val = day_stat["first_tap"].strftime("%H:%M") if c_count >= 1 and day_stat["first_tap"] else "-"
            last_val  = "-"
            hours_std_val = "-"
            hours_ot_val  = "-"
            late_val      = "-"
            early_val     = "-"

            if c_count >= 2 and day_stat["first_tap"] != day_stat["last_tap"]:
                last_val = day_stat["last_tap"].strftime("%H:%M")
                _, hours_std, hours_ot, min_late, min_early = compute_day_stats(
                    day_stat["first_tap"], day_stat["last_tap"],
                    d, emp_dept_raw, row_shift
                )
                hours_std_val = hours_std
                hours_ot_val  = hours_ot if hours_ot > 0 else 0
                late_val      = min_late
                early_val     = min_early

            flat_row = [
                emp_id, emp_name, emp_dept, emp_group, shift_text,
                d.strftime("%d/%m/%Y"),
                first_val, last_val,
                hours_std_val, hours_ot_val,
                late_val, early_val
            ]
            ws2.append(flat_row)
            last_ws2_row = ws2.max_row
            for col_idx, cell in enumerate(ws2[last_ws2_row], start=1):
                cell.border = border_style
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Apply AutoFilter to sheet 2
    ws2.auto_filter.ref = f"A1:L{ws2.max_row}"
    # ──────────────────────────────────────────────────────────────────
        
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    wb.close()
    
    return FileResponse(
        path=temp_file.name, 
        filename=f"ChamCong_{start_date}_{end_date}.xlsx", 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(os.remove, temp_file.name)
    )

app.mount("/", StaticFiles(directory="static", html=True), name="static")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Run the Time Attendance System API.")
    parser.add_argument("--host", type=str, default="0.0.0.0", help="Host address to bind to")
    parser.add_argument("--port", type=int, default=8001, help="Port to listen on")
    
    args = parser.parse_args()
    
    uvicorn.run(app, host=args.host, port=args.port)
