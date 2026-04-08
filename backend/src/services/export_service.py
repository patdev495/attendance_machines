import os
import threading
import tempfile
from datetime import date, datetime, timedelta
from sqlalchemy import text, func, Date, case
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from database import AttendanceLog, EmployeeMetadata, ShiftRule, SessionLocal
from utils.stats_utils import compute_day_stats

export_status = {
    "is_running": False,
    "progress": 0,
    "total": 0,
    "current_step": "",
    "cancel_requested": False,
    "filename": None,
    "error": None
}
export_lock = threading.Lock()

def run_export_task(start_date: date, end_date: date, view_mode: str):
    global export_status
    
    db = SessionLocal()
    try:
        with export_lock:
            export_status.update({
                "is_running": True, "progress": 0, "total": 0, 
                "current_step": "Fetching data...", "cancel_requested": False, 
                "filename": None, "error": None
            })

        # Pre-fetch shift rules for extreme performance
        rules_pool = db.query(ShiftRule).all()

        # Base query for attendance aggregation
        base_calc_sub = db.query(
            AttendanceLog.id,
            AttendanceLog.attendance_time,
            AttendanceLog.employee_id,
            EmployeeMetadata.shift,
            case(
                (EmployeeMetadata.shift == text("'D'"), func.cast(func.dateadd(text("hour"), text("-10"), AttendanceLog.attendance_time), Date)),
                else_=func.cast(func.dateadd(text("hour"), text("-4"), AttendanceLog.attendance_time), Date)
            ).label("work_date")
        ).outerjoin(EmployeeMetadata, AttendanceLog.employee_id == EmployeeMetadata.employee_id).subquery()
        
        query = db.query(
            base_calc_sub.c.employee_id,
            base_calc_sub.c.work_date,
            func.min(base_calc_sub.c.attendance_time).label("first_tap"),
            func.max(base_calc_sub.c.attendance_time).label("last_tap"),
            func.count(base_calc_sub.c.id).label("tap_count"),
            base_calc_sub.c.shift
        )
        query = query.filter(base_calc_sub.c.work_date >= start_date)
        query = query.filter(base_calc_sub.c.work_date <= end_date)
        query = query.group_by(
            base_calc_sub.c.employee_id, 
            base_calc_sub.c.work_date, 
            base_calc_sub.c.shift
        )
        
        results = query.all()
        if not results:
             with export_lock:
                 export_status.update({"is_running": False, "error": "No data found for this range."})
             return

        # Extract set of employee IDs from results
        all_emp_ids = sorted(list(set(row.employee_id for row in results)))
        emp_meta_query = db.query(EmployeeMetadata).filter(EmployeeMetadata.employee_id.in_(all_emp_ids)).all()
        emp_meta = {m.employee_id: m for m in emp_meta_query}
        
        # Pre-define styles for high performance
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        alignment_style = Alignment(horizontal="center", vertical="center", wrap_text=True)
        white_font = Font(color="FFFFFF")
        bold_font = Font(bold=True)

        # Process results into a dictionary with pre-calculated stats
        processed_data = {}
        with export_lock: export_status["current_step"] = "Preparing attendance statistics..."
        
        for row in results:
            emp_id = row.employee_id
            w_date = row.work_date
            if emp_id not in processed_data:
                processed_data[emp_id] = {"general_shift": row.shift, "days": {}}
            
            first_val, last_val, std, ot, late, early = "-", "-", "-", "-", "-", "-"
            if row.tap_count >= 1: first_val = row.first_tap.strftime("%H:%M")
            if row.tap_count >= 2:
                last_val = row.last_tap.strftime("%H:%M")
                emp_m = emp_meta.get(emp_id)
                _, std, ot, late, early = compute_day_stats(
                    row.first_tap, 
                    row.last_tap, 
                    w_date, 
                    emp_m.department if emp_m else None, 
                    row.shift,
                    rules_pool=rules_pool
                )
            
            processed_data[emp_id]["days"][w_date] = {
                "first": first_val, "last": last_val, 
                "std": std, "ot": ot, "late": late, "early": early,
                "shift": row.shift
            }
            
        sorted_emp_ids = sorted(processed_data.keys())
            
        dates_list = []
        curr = start_date
        while curr <= end_date:
            dates_list.append(curr)
            curr += timedelta(days=1)
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Export"
        headers = ["Mã nhân viên", "Tên nhân viên", "Phòng ban", "Nhóm", "Ngày vào làm", "Ca làm việc", "Ghi chú"]
        for d in dates_list: headers.append(f"{d.day}/{d.month}")
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = alignment_style
            cell.border = border_style
            cell.fill = header_fill

        with export_lock:
            export_status["total"] = len(sorted_emp_ids)
            export_status["current_step"] = "Generating Sheet 1..."

        current_row = 2
        num_rows = 6 if view_mode == "both" else 4
        
        for idx, emp_id in enumerate(sorted_emp_ids):
            if export_status["cancel_requested"]: 
                with export_lock: export_status["is_running"] = False
                return
            
            if idx % 5 == 0:
                with export_lock:
                    export_status["progress"] = int((idx / len(sorted_emp_ids)) * 50) 

            emp_data_info = processed_data[emp_id]
            emp_data = emp_data_info["days"]
            shift_val = emp_data_info["general_shift"]
            emp_m = emp_meta.get(emp_id)
            emp_name = emp_m.emp_name if emp_m and emp_m.emp_name else emp_id
            emp_dept = emp_m.department if emp_m and emp_m.department else "-"
            emp_group = emp_m.group if emp_m and emp_m.group else "-"
            emp_start = emp_m.start_date.strftime("%d/%m/%Y") if emp_m and emp_m.start_date else "-"
            shift_text = shift_val if shift_val else "-"

            for r_offset in range(num_rows):
                r_idx = current_row + r_offset
                is_first = (r_offset == 0)
                
                # Optimized cell assignment with inline styling
                vals = [emp_id, emp_name, emp_dept, emp_group, emp_start, shift_text]
                for c_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = border_style
                    cell.alignment = alignment_style
                    if not is_first:
                        cell.font = white_font
                
                # Indicator column
                if view_mode == "time": all_indicators = ["Giờ In", "Giờ Out", "Đi muộn (phút)", "Về sớm (phút)"]
                elif view_mode == "hours": all_indicators = ["Giờ Công", "Tăng Ca", "Đi muộn (phút)", "Về sớm (phút)"]
                else: all_indicators = ["Giờ In", "Giờ Out", "Giờ Công", "Tăng Ca", "Đi muộn (phút)", "Về sớm (phút)"]
                
                cell_ind = ws.cell(row=r_idx, column=7, value=all_indicators[r_offset])
                cell_ind.border = border_style
                cell_ind.alignment = alignment_style

            for date_idx, d in enumerate(dates_list):
                col = date_idx + 8
                if d not in emp_data:
                    for i in range(num_rows): 
                        cell = ws.cell(row=current_row+i, column=col, value="-")
                        cell.border = border_style
                        cell.alignment = alignment_style
                else:
                    ds = emp_data[d]
                    def fmt_ot(v):
                        return v if (isinstance(v, (int, float)) and v > 0) else "-"

                    if view_mode=="time":
                        vals = [ds["first"], ds["last"], ds["late"], ds["early"]]
                    elif view_mode=="hours":
                        vals = [ds["std"], fmt_ot(ds["ot"]), ds["late"], ds["early"]]
                    else:
                        vals = [ds["first"], ds["last"], ds["std"], fmt_ot(ds["ot"]), ds["late"], ds["early"]]
                    
                    for i, v in enumerate(vals): 
                        cell = ws.cell(row=current_row+i, column=col, value=v)
                        cell.border = border_style
                        cell.alignment = alignment_style
            current_row += num_rows

        # Sheet 2 - Information Detail
        with export_lock: export_status["current_step"] = "Generating Sheet 2..."
        ws2 = wb.create_sheet(title="Thông tin chi tiết")
        ws2.append(["Mã NV", "Tên nhân viên", "Phòng ban", "Nhóm", "Ca", "Ngày", "Giờ In", "Giờ Out", "Giờ Công", "Tăng Ca", "Đi muộn (phút)", "Về sớm (phút)"])
        
        for cell in ws2[1]:
            cell.font = bold_font
            cell.alignment = alignment_style
            cell.border = border_style
            cell.fill = header_fill
        
        ws2_row = 2
        for idx, emp_id in enumerate(sorted_emp_ids):
            if export_status["cancel_requested"]: 
                with export_lock: export_status["is_running"] = False
                return
            
            if idx % 10 == 0:
                with export_lock: export_status["progress"] = 50 + int((idx / len(sorted_emp_ids)) * 50)
            
            emp_info = processed_data[emp_id]
            emp_m = emp_meta.get(emp_id)
            for d in dates_list:
                if d not in emp_info["days"]: continue
                ds = emp_info["days"][d]
                
                ot = ds["ot"]
                display_ot = ot if (isinstance(ot, (int, float)) and ot > 0) else 0
                row_data = [
                    emp_id, emp_m.emp_name if emp_m else emp_id, emp_m.department if emp_m else "-", 
                    emp_m.group if emp_m else "-", ds["shift"] or "N", d.strftime("%d/%m/%Y"), 
                    ds["first"], ds["last"], ds["std"], display_ot, ds["late"], ds["early"]
                ]
                
                for c_idx, val in enumerate(row_data, 1):
                    cell = ws2.cell(row=ws2_row, column=c_idx, value=val)
                    cell.border = border_style
                    cell.alignment = alignment_style
                
                ws2_row += 1

        # Optimized Auto-width (only once per column)
        for sheet in [ws, ws2]:
            for col_idx in range(1, sheet.max_column + 1):
                column_letter = get_column_letter(col_idx)
                # Sample a few rows for width to avoid full iteration if possible, 
                # but for accuracy we still check header and first 5 data rows
                max_length = 0
                for row_idx in range(1, min(sheet.max_row, 20) + 1):
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if val:
                        length = max(len(str(line)) for line in str(val).split('\n'))
                        if length > max_length: max_length = length
                
                adjusted_width = (max_length + 4) * 1.1 
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 60)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        with export_lock:
            export_status.update({"is_running": False, "progress": 100, "filename": tmp.name, "current_step": "Ready for download"})
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        with export_lock:
            export_status.update({"is_running": False, "error": str(e)})
    finally:
        db.close()
