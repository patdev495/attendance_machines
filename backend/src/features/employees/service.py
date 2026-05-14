from sqlalchemy.orm import Session
from database import EmployeeLocalRegistry, EmployeeMetadata, AttendanceLog, SessionLocal
from config import config, DEMO_MODE
from shared.hardware import get_machine_list

if not DEMO_MODE:
    from features.machines.service import get_users_from_machine, delete_user_from_machine, update_user_name_on_machine
from concurrent.futures import ThreadPoolExecutor
import concurrent.futures
import logging
import io
import openpyxl
from sqlalchemy import func, Integer
from compat import safe_ilike

logger = logging.getLogger(__name__)

def update_registry(db: Session):
    """
    Update EmployeeLocalRegistry from 3 sources:
    1. Excel (EmployeeMetadata)
    2. Machines (get_users_from_machine)
    3. Logs (AttendanceLogs)
    """
    try:
        # 1. Gather current sets
        excel_users = {emp.employee_id: emp for emp in db.query(EmployeeMetadata).all()}
        
        machine_users = set()
        if not DEMO_MODE:
            ips = get_machine_list()
            for ip in ips:
                users, status = get_users_from_machine(ip)
                if status == "Success":
                    for u in users:
                        machine_users.add(str(u.get('user_id')))
                    
        log_users = {emp_id for (emp_id,) in db.query(AttendanceLog.employee_id).distinct().all()}
        
        all_active_ids = set(excel_users.keys()) | machine_users | log_users
        
        # 2. Update existing and delete stale
        existing_registry = db.query(EmployeeLocalRegistry).all()
        for reg in existing_registry:
            if reg.employee_id not in all_active_ids:
                db.delete(reg)
            else:
                # Update status based on precedence
                if reg.employee_id in excel_users:
                    emp = excel_users[reg.employee_id]
                    reg.emp_name = emp.emp_name
                    reg.department = emp.department
                    reg.group_name = emp.group
                    reg.start_date = emp.start_date
                    reg.shift = emp.shift
                    reg.source_status = 'excel_synced'
                elif reg.employee_id in machine_users:
                    reg.source_status = 'machine_only'
                elif reg.employee_id in log_users:
                    reg.source_status = 'log_only'
                    
        db.commit()
        
        # 3. Add new entries
        existing_ids = {reg.employee_id for reg in db.query(EmployeeLocalRegistry.employee_id).all()}
        new_ids = all_active_ids - existing_ids
        
        for emp_id in new_ids:
            if emp_id in excel_users:
                emp = excel_users[emp_id]
                new_reg = EmployeeLocalRegistry(
                    employee_id=emp_id,
                    emp_name=emp.emp_name,
                    department=emp.department,
                    group_name=emp.group,
                    start_date=emp.start_date,
                    shift=emp.shift,
                    source_status='excel_synced'
                )
            elif emp_id in machine_users:
                new_reg = EmployeeLocalRegistry(employee_id=emp_id, source_status='machine_only')
            else:
                new_reg = EmployeeLocalRegistry(employee_id=emp_id, source_status='log_only')
                
            db.add(new_reg)
            
        db.commit()
    except Exception as e:
        logger.error(f"Error updating registry: {e}")
        db.rollback()
        raise

def delete_user_from_hardware(employee_id: str):
    """
    Deletes user ONLY from hardware, does not delete from DB.
    Uses ThreadPoolExecutor for parallel processing.
    """
    ips = get_machine_list()
    results = {}
    
    with ThreadPoolExecutor(max_workers=max(1, len(ips))) as executor:
        future_to_ip = {executor.submit(delete_user_from_machine, ip, employee_id): ip for ip in ips}
        for future in concurrent.futures.as_completed(future_to_ip):
            ip = future_to_ip[future]
            try:
                result = future.result()
                results[ip] = result
            except Exception as e:
                results[ip] = f"Error: {str(e)}"
    return results

def update_employee_info(employee_id: str, db_name: str, db: Session):
    """
    Updates user info in EmployeeLocalRegistry and EmployeeMetadata.
    """
    # 1. Update EmployeeLocalRegistry
    registry_entry = db.query(EmployeeLocalRegistry).filter(EmployeeLocalRegistry.employee_id == employee_id).first()
    if registry_entry:
        registry_entry.emp_name = db_name
        
    # 2. Update EmployeeMetadata
    meta_entry = db.query(EmployeeMetadata).filter(EmployeeMetadata.employee_id == employee_id).first()
    if meta_entry:
        meta_entry.emp_name = db_name
        
    db.commit()
    
    return {"status": "success", "message": "Updated in DB"}

def export_employees_to_excel(db: Session, search: str = None, source_status: str = None, privilege: int = None):
    query = db.query(EmployeeLocalRegistry)
    
    if search:
        search = search.strip()
        found_ids = db.query(EmployeeLocalRegistry.employee_id).filter(
            EmployeeLocalRegistry.employee_id.ilike(f"%{search}%") |
            EmployeeLocalRegistry.full_emp_id.ilike(f"%{search}%") |
            safe_ilike(EmployeeLocalRegistry.emp_name, f"%{search}%")
        ).all()
        target_ids = {r[0] for r in found_ids} | {search}
        query = query.filter(func.ltrim(func.rtrim(EmployeeLocalRegistry.employee_id)).in_(list(target_ids)))
        
    if source_status:
        query = query.filter(EmployeeLocalRegistry.source_status == source_status)
        
    if privilege is not None:
        query = query.filter(EmployeeLocalRegistry.privilege == privilege)
        
    query = query.order_by(func.cast(EmployeeLocalRegistry.employee_id, Integer).asc())
    employees = query.all()
    
    # Lấy thông tin chấm công gần nhất cho mỗi nhân viên
    latest_logs = db.query(
        AttendanceLog.employee_id, 
        func.max(AttendanceLog.attendance_time).label('last_time')
    ).group_by(AttendanceLog.employee_id).all()
    latest_log_dict = {log.employee_id: log.last_time for log in latest_logs}
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    headers = [
        "Mã NV (Máy CC)", "Mã NV (Đầy đủ)", "Họ và Tên", "Phòng Ban", 
        "Nhóm / Chuyền", "Ngày vào làm", "Ca làm việc", "Nguồn dữ liệu", "Lần chấm công gần nhất"
    ]
    ws.append(headers)
    
    for emp in employees:
        last_time = latest_log_dict.get(emp.employee_id)
        ws.append([
            emp.employee_id,
            emp.full_emp_id or "",
            emp.emp_name or "",
            emp.department or "",
            emp.group_name or "",
            (emp.start_date if emp.start_date and emp.start_date.year > 1970 else ""),
            emp.shift or "",
            emp.source_status or "",
            last_time if last_time else ""
        ])
        
        # Định dạng lại cột Ngày (cột 6) và cột Thời gian (cột 9) theo chuẩn Date/Time của Excel
        current_row = ws.max_row
        if emp.start_date:
            ws.cell(row=current_row, column=6).number_format = 'yyyy-mm-dd'
        if last_time:
            ws.cell(row=current_row, column=9).number_format = 'yyyy-mm-dd hh:mm:ss'
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col_letter].width = adjusted_width
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
