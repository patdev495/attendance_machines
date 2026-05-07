import os
import threading
import tempfile
from datetime import date, datetime, timedelta, time
from sqlalchemy import text, func, Date, Time, case, literal_column
from sqlalchemy.orm import aliased
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from database import AttendanceLog, EmployeeLocalRegistry, EmployeeDailyShifts, ShiftDefinition, SessionLocal

from utils.stats_utils import compute_day_stats, parse_shift_window, FULL_DAY_LEAVE_CODES, determine_missing_tap

export_status = {
    "is_running": False,
    "progress": 0,
    "total": 0,
    "current_step": "",
    "cancel_requested": False,
    "filename": None,
    "error": None
}
export_lock = threading.Lock()

def _get_shift_meta(shift_code, rules_pool):
    """Return (shift_category, is_night_shift) for a given shift_code.

    shift_category values: 'NORMAL' | 'HOLIDAY' | 'ROTATION'
    Defaults to ('NORMAL', False) when the code is not in rules_pool.
    """
    if rules_pool and shift_code:
        code = shift_code.strip().upper()
        for r in rules_pool:
            if r.shift_code == code:
                cat = (r.shift_category or "NORMAL").upper()
                return cat, bool(r.is_night_shift)
    return "NORMAL", False

def run_export_task(start_date: date, end_date: date, view_mode: str):
    global export_status
    
    db = SessionLocal()
    try:
        with export_lock:
            export_status.update({
                "is_running": True, "progress": 0, "total": 0, 
                "current_step": "Fetching data...", "cancel_requested": False, 
                "filename": None, "error": None
            })

        # Pre-fetch shift rules for extreme performance
        rules_pool = db.query(ShiftDefinition).all()


        # Phase 13: Use 09:00 AM anchor for work_date (consistent with router.py)
        # Phase 15 (export): Use the same shift-aware work_date logic as the daily summary router.
        # The old -12h anchor was only based on EmployeeLocalRegistry.shift, which is often empty
        # when the employee's shift is set only via daily shift codes (EmployeeDailyShifts).
        # This fix aligns export grouping with the tab view.
        today_shift = aliased(EmployeeDailyShifts)
        yest_shift  = aliased(EmployeeDailyShifts)

        base_calc_sub = db.query(
            AttendanceLog.id,
            AttendanceLog.attendance_time,
            AttendanceLog.employee_id,
            EmployeeLocalRegistry.shift,
            EmployeeLocalRegistry.department,
            EmployeeLocalRegistry.full_emp_id,
            case(
                # Priority 1: Today is D-like + tap before noon → belongs to yesterday
                (
                    (func.coalesce(today_shift.shift_code, EmployeeLocalRegistry.shift).like("%D%")) &
                    (func.cast(AttendanceLog.attendance_time, Time) < time(12, 0)),
                    func.cast(func.dateadd(text("day"), text("-1"), AttendanceLog.attendance_time), Date)
                ),
                # Priority 2: Today is D-like + tap ≥ 18:00 → belongs to today
                (
                    (func.coalesce(today_shift.shift_code, EmployeeLocalRegistry.shift).like("%D%")) &
                    (func.cast(AttendanceLog.attendance_time, Time) >= time(18, 0)),
                    func.cast(AttendanceLog.attendance_time, Date)
                ),
                # Priority 3: Yesterday was D-like + early morning tap → belongs to yesterday
                (
                    (func.coalesce(yest_shift.shift_code, EmployeeLocalRegistry.shift).like("%D%")) &
                    (func.cast(AttendanceLog.attendance_time, Time) < time(10, 0)),
                    func.cast(func.dateadd(text("day"), text("-1"), AttendanceLog.attendance_time), Date)
                ),
                # Default: -3h anchor (day shifts)
                else_=func.cast(func.dateadd(text("hour"), text("-3"), AttendanceLog.attendance_time), Date)
            ).label("work_date")
        ).outerjoin(
            EmployeeLocalRegistry,
            func.ltrim(func.rtrim(AttendanceLog.employee_id)) == func.ltrim(func.rtrim(EmployeeLocalRegistry.employee_id))
        ).outerjoin(
            today_shift,
            (func.ltrim(func.rtrim(AttendanceLog.employee_id)) == func.ltrim(func.rtrim(today_shift.employee_id))) &
            (func.cast(AttendanceLog.attendance_time, Date) == today_shift.work_date)
        ).outerjoin(
            yest_shift,
            (func.ltrim(func.rtrim(AttendanceLog.employee_id)) == func.ltrim(func.rtrim(yest_shift.employee_id))) &
            (func.cast(func.dateadd(text("day"), text("-1"), AttendanceLog.attendance_time), Date) == yest_shift.work_date)
        ).subquery()
        
        # Phase 13: JOIN with EmployeeDailyShifts
        query = db.query(
            base_calc_sub.c.employee_id,
            base_calc_sub.c.full_emp_id,
            base_calc_sub.c.work_date,
            func.min(base_calc_sub.c.attendance_time).label("first_tap"),
            func.max(base_calc_sub.c.attendance_time).label("last_tap"),
            func.count(base_calc_sub.c.id).label("tap_count"),
            # Use daily shift code first, fallback to employee-level shift
            func.coalesce(
                func.max(EmployeeDailyShifts.shift_code),
                base_calc_sub.c.shift
            ).label("shift"),
            base_calc_sub.c.department,
            func.max(EmployeeDailyShifts.shift_code).label("daily_shift_code")
        ).outerjoin(
            EmployeeDailyShifts,
            (func.ltrim(func.rtrim(base_calc_sub.c.employee_id)) == EmployeeDailyShifts.employee_id) &
            (base_calc_sub.c.work_date == EmployeeDailyShifts.work_date)
        )
        query = query.filter(base_calc_sub.c.work_date >= start_date)
        query = query.filter(base_calc_sub.c.work_date <= end_date)
        query = query.group_by(
            base_calc_sub.c.employee_id, 
            base_calc_sub.c.full_emp_id,
            base_calc_sub.c.work_date, 
            base_calc_sub.c.shift, 
            base_calc_sub.c.department
        ).order_by(base_calc_sub.c.work_date, base_calc_sub.c.employee_id)
        
        results = query.all()
        if not results:
             with export_lock:
                 export_status.update({"is_running": False, "error": "No data found for this range."})
             return

        # Also load ALL daily shifts for the date range (for days without attendance logs)
        all_daily_shifts = db.query(EmployeeDailyShifts).filter(
            EmployeeDailyShifts.work_date >= start_date,
            EmployeeDailyShifts.work_date <= end_date
        ).all()
        # Build a lookup: {(emp_id, work_date): shift_code}
        daily_shift_lookup = {}
        for ds in all_daily_shifts:
            daily_shift_lookup[(ds.employee_id, ds.work_date)] = ds.shift_code

        # Extract set of employee IDs from results
        all_emp_ids = sorted(list(set(row.employee_id for row in results)))
        # Also include employees that have daily shifts but no attendance logs
        daily_shift_emp_ids = set(ds.employee_id for ds in all_daily_shifts)
        all_emp_ids = sorted(list(set(all_emp_ids) | daily_shift_emp_ids))

        # Phase 13: Batch fetch to avoid SQL Server 2100 parameter limit
        emp_meta = {}
        for i in range(0, len(all_emp_ids), 1000):
            batch = all_emp_ids[i:i+1000]
            batch_query = db.query(EmployeeLocalRegistry).filter(EmployeeLocalRegistry.employee_id.in_(batch)).all()
            for m in batch_query:
                emp_meta[m.employee_id] = m
        
        # Pre-define styles
        border_style = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        alignment_style = Alignment(horizontal="center", vertical="center", wrap_text=True)
        white_font = Font(color="FFFFFF")
        bold_font = Font(bold=True)

        processed_data = {}
        with export_lock: export_status["current_step"] = "Preparing attendance statistics..."
        
        for row in results:
            emp_id = row.employee_id
            w_date = row.work_date
            if emp_id not in processed_data:
                processed_data[emp_id] = {"general_shift": row.shift, "days": {}}
            
            # Get effective shift code (daily override or fallback)
            daily_code = getattr(row, 'daily_shift_code', None)
            full_emp_id = getattr(row, 'full_emp_id', None)
            
            # Phase 13: Map to NA if not in definitions table
            valid_codes = {r.shift_code for r in rules_pool} if rules_pool else set()
            effective_shift_raw = daily_code or row.shift or "N"
            
            if not effective_shift_raw or effective_shift_raw.strip().upper() not in valid_codes:
                shift_code_display = "NA"
                effective_shift = "N"
            else:
                shift_code_display = effective_shift_raw.strip().upper()
                effective_shift = shift_code_display
            
            # Initialization for row record
            first_val, last_val, std, ot, late, early = "-", "-", "-", "-", "-", "-"
            night_subsidy = 0.0
            hp = hr = ho = ht = hc = hk = 0.0
            std_expected = 8.0
            workday_base = 8.0
            note = ""

            # Check for full-day leave
            if effective_shift.strip().upper() in FULL_DAY_LEAVE_CODES:
                std = 0
                ot = 0
                late = 0
                early = 0
                note = effective_shift.strip().upper()
                if row.tap_count >= 1: first_val = row.first_tap.strftime("%H:%M")
                if row.tap_count >= 2: last_val = row.last_tap.strftime("%H:%M")
                # Phase 15: Get standard hours and leave types for leave days
                emp_m = emp_meta.get(emp_id)
                dept = getattr(row, 'department', None) or (emp_m.department if emp_m else None)
                win = parse_shift_window(effective_shift, dept, rules_pool=rules_pool)
                std_expected = win.get('standard_hours', 8.0)
                workday_base = win.get('workday_base', 8.0)
                hp = win.get('leave_hours_p', 0.0)
                hr = win.get('leave_hours_r', 0.0)
                ho = win.get('leave_hours_o', 0.0)
                ht = win.get('leave_hours_t', 0.0)
                hc = win.get('leave_hours_c', 0.0)
                hk = win.get('leave_hours_k', 0.0)
            elif row.tap_count >= 1:
                first_val = row.first_tap.strftime("%H:%M")
                last_val = row.last_tap.strftime("%H:%M")
                emp_m = emp_meta.get(emp_id)
                department = getattr(row, 'department', None) or (emp_m.department if emp_m else None)
                res_work, std, ot, late, early, hp, hr, ho, ht, hc, hk, night_subsidy, std_expected, workday_base = compute_day_stats(
                    row.first_tap, 
                    row.last_tap, 
                    w_date, 
                    department, 
                    effective_shift,
                    rules_pool=rules_pool
                )
                # If only 1 tap, OR multiple taps but they are within 5 minutes (duplicate taps)
                if row.tap_count == 1 or (row.last_tap - row.first_tap).total_seconds() < 300:
                    note = determine_missing_tap(row.first_tap, w_date, effective_shift, department, rules_pool)
                    # Single tap: suppress the metric we cannot know
                    if note == "Missing Check-out":
                        early = 0   # don't know when they left
                        last_val = "-"
                    elif note == "Missing Check-in":
                        late = 0    # don't know when they arrived
                        first_val = "-"
                    
                    # If there's leave hours, we might want to suppress the "Missing" note 
                    # but we ALWAYS keep the '-' in first_val/last_val as requested.
                    if (hp or hr or ho or ht or hc or hk):
                        note = "" # Leave note will be handled by the shift code or general notes

            
            # ── Classify std/ot by shift_category (NORMAL / HOLIDAY / ROTATION)
            #    and day vs night shift. Logic unchanged — only routing results.
            _cat, _is_night = _get_shift_meta(effective_shift, rules_pool)
            _std_num = std if isinstance(std, (int, float)) else 0.0
            _ot_num  = ot  if isinstance(ot,  (int, float)) else 0.0
            work_normal = work_holiday = work_rotation = 0.0
            ot_normal_day = ot_holiday_day = ot_rotation_day = 0.0
            ot_normal_night = ot_holiday_night = ot_rotation_night = 0.0
            if _cat == "HOLIDAY":
                work_holiday = _std_num
                if _is_night: ot_holiday_night = _ot_num
                else:         ot_holiday_day   = _ot_num
            elif _cat == "ROTATION":
                work_rotation = _std_num
                if _is_night: ot_rotation_night = _ot_num
                else:         ot_rotation_day   = _ot_num
            else:  # NORMAL (default)
                work_normal = _std_num
                if _is_night: ot_normal_night = _ot_num
                else:         ot_normal_day   = _ot_num
 
            processed_data[emp_id]["days"][w_date] = {
                "first": first_val, "last": last_val, 
                "std": std, "ot": ot, "late": late, "early": early,
                "shift": shift_code_display,
                "shift_code": shift_code_display,
                "note": note,
                "night_subsidy": night_subsidy,
                "std_expected": std_expected,
                "workday_base": workday_base,
                "hp": hp, "hr": hr, "ho": ho, "ht": ht, "hc": hc, "hk": hk,
                # 9 cột phân loại
                "work_normal": work_normal, "work_holiday": work_holiday, "work_rotation": work_rotation,
                "ot_normal_day": ot_normal_day, "ot_holiday_day": ot_holiday_day, "ot_rotation_day": ot_rotation_day,
                "ot_normal_night": ot_normal_night, "ot_holiday_night": ot_holiday_night, "ot_rotation_night": ot_rotation_night,
            }
            if full_emp_id:
                processed_data[emp_id]["full_id"] = full_emp_id

        # Phase 13: Add days with leave codes but no attendance logs
        for (emp_id, w_date), shift_code in daily_shift_lookup.items():
            if emp_id not in processed_data:
                processed_data[emp_id] = {"general_shift": shift_code, "days": {}}
            if w_date not in processed_data[emp_id]["days"]:
                code_upper = shift_code.strip().upper()
                if code_upper in FULL_DAY_LEAVE_CODES:
                    processed_data[emp_id]["days"][w_date] = {
                        "first": "-", "last": "-",
                        "std": 0, "ot": 0, "late": 0, "early": 0,
                        "shift": code_upper,
                        "shift_code": code_upper,
                        # Leave days: all category buckets = 0
                        "work_normal": 0, "work_holiday": 0, "work_rotation": 0,
                        "ot_normal_day": 0, "ot_holiday_day": 0, "ot_rotation_day": 0,
                        "ot_normal_night": 0, "ot_holiday_night": 0, "ot_rotation_night": 0,
                    }
            
        sorted_emp_ids = sorted(processed_data.keys())
            
        dates_list = []
        curr = start_date
        while curr <= end_date:
            dates_list.append(curr)
            curr += timedelta(days=1)
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Export"
        # Column 7 renamed to "Chỉ số", and real "Ghi chú" added at the end
        headers = ["Mã máy", "Mã công ty", "Tên nhân viên", "Phòng ban", "Nhóm", "Ngày vào làm", "Chỉ số"]
        for d in dates_list: headers.append(f"{d.day}/{d.month}")
        headers.append("Ghi chú")
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = alignment_style
            cell.border = border_style
            cell.fill = header_fill

        with export_lock:
            export_status["total"] = len(sorted_emp_ids)
            export_status["current_step"] = "Generating Sheet 1..."

        current_row = 2
        # Phase 13: Add "Mã công" row to the indicator set
        if view_mode == "time":
            all_indicators = ["Giờ In", "Giờ Out", "Đi muộn (phút)", "Về sớm (phút)", "Mã công"]
        elif view_mode == "hours":
            all_indicators = ["Giờ Công", "Tăng Ca", "Đi muộn (phút)", "Về sớm (phút)", "Mã công"]
        else:
            all_indicators = ["Giờ In", "Giờ Out", "Giờ Công", "Tăng Ca", "Đi muộn (phút)", "Về sớm (phút)", "Mã công"]
        
        num_rows = len(all_indicators)
        
        for idx, emp_id in enumerate(sorted_emp_ids):
            if export_status["cancel_requested"]: 
                with export_lock: export_status["is_running"] = False
                return
            
            if idx % 5 == 0:
                with export_lock:
                    export_status["progress"] = int((idx / len(sorted_emp_ids)) * 50) 

            emp_data_info = processed_data[emp_id]
            emp_data = emp_data_info["days"]
            shift_val = emp_data_info["general_shift"]
            emp_m = emp_meta.get(emp_id)
            emp_name = emp_m.emp_name if emp_m and emp_m.emp_name else emp_id
            emp_dept = emp_m.department if emp_m and emp_m.department else "-"
            emp_group = emp_m.group_name if emp_m and emp_m.group_name else "-"
            emp_start = emp_m.start_date.strftime("%d/%m/%Y") if emp_m and emp_m.start_date else "-"
            shift_text = shift_val if shift_val else "-"

            for r_offset in range(num_rows):
                r_idx = current_row + r_offset
                is_first = (r_offset == 0)
                
                full_id_val = emp_data_info.get("full_id") or "-"
                vals = [emp_id, full_id_val, emp_name, emp_dept, emp_group, emp_start]
                for c_idx, val in enumerate(vals, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = border_style
                    cell.alignment = alignment_style
                    if not is_first:
                        cell.font = white_font
                
                cell_ind = ws.cell(row=r_idx, column=7, value=all_indicators[r_offset])
                cell_ind.border = border_style
                cell_ind.alignment = alignment_style

            # Fill in notes in the last column of the first row of each employee block
            notes_str = ""
            for d in dates_list:
                if d in emp_data and emp_data[d].get("note"):
                    if notes_str: notes_str += "; "
                    notes_str += f"{d.day}/{d.month}: {emp_data[d]['note']}"
            
            notes_col = len(dates_list) + 8
            ws.cell(row=current_row, column=notes_col, value=notes_str if notes_str else "-").border = border_style
            for i in range(1, num_rows):
                ws.cell(row=current_row+i, column=notes_col, value="").border = border_style

            for date_idx, d in enumerate(dates_list):
                col = date_idx + 8 # Date columns start after "Chỉ số" (Column 7)
                if d not in emp_data:
                    for i in range(num_rows): 
                        cell = ws.cell(row=current_row+i, column=col, value="-")
                        cell.border = border_style
                        cell.alignment = alignment_style
                else:
                    ds = emp_data[d]
                    def fmt_ot(v):
                        return v if (isinstance(v, (int, float)) and v > 0) else "-"

                    if view_mode=="time":
                        vals = [ds["first"], ds["last"], ds["late"], ds["early"], ds.get("shift_code", "-")]
                    elif view_mode=="hours":
                        vals = [ds["std"], fmt_ot(ds["ot"]), ds["late"], ds["early"], ds.get("shift_code", "-")]
                    else:
                        vals = [ds["first"], ds["last"], ds["std"], fmt_ot(ds["ot"]), ds["late"], ds["early"], ds.get("shift_code", "-")]
                    
                    for i, v in enumerate(vals): 
                        cell = ws.cell(row=current_row+i, column=col, value=v)
                        cell.border = border_style
                        cell.alignment = alignment_style
            current_row += num_rows

        with export_lock: export_status["current_step"] = "Generating Sheet 2..."
        ws2 = wb.create_sheet(title="Thông tin chi tiết")
        # Added "Ngày vào làm" to Sheet 2
        ws2.append([
            "Mã máy", "Mã công ty", "Tên nhân viên", "Phòng ban", "Nhóm", "Ngày vào làm",
            "Ngày", "Giờ In", "Giờ Out", "Giờ Công", "Tăng Ca", "Đi muộn (phút)", "Về sớm (phút)", "Mã công",
            "Công tiêu chuẩn",
            "Giờ trợ cấp ca đêm",
            # 9 cột phân loại theo shift_category
            "Giờ ngày thường", "Giờ ngày nghỉ lễ", "Giờ ngày nghỉ luân phiên",
            "TC thường - ca ngày", "TC lễ - ca ngày", "TC luân phiên - ca ngày",
            "TC thường - ca đêm", "TC lễ - ca đêm", "TC luân phiên - ca đêm",
        ])
        
        for cell in ws2[1]:
            cell.font = bold_font
            cell.alignment = alignment_style
            cell.border = border_style
            cell.fill = header_fill
        
        ws2_row = 2
        for idx, emp_id in enumerate(sorted_emp_ids):
            if export_status["cancel_requested"]: 
                with export_lock: export_status["is_running"] = False
                return
            
            if idx % 10 == 0:
                with export_lock: export_status["progress"] = 50 + int((idx / len(sorted_emp_ids)) * 50)
            
            emp_info = processed_data[emp_id]
            emp_m = emp_meta.get(emp_id)
            for d in dates_list:
                if d not in emp_info["days"]: continue
                ds = emp_info["days"][d]
                
                ot = ds["ot"]
                display_ot = ot if (isinstance(ot, (int, float)) and ot > 0) else 0
                shift_code_val = ds.get("shift_code", ds.get("shift", "N"))
                full_id_val = emp_info.get("full_id") or "-"
                hired_date_val = emp_m.start_date.strftime("%d/%m/%Y") if emp_m and emp_m.start_date else "-"
                row_data = [
                    emp_id, full_id_val, emp_m.emp_name if emp_m else emp_id, 
                    emp_m.department if emp_m else "-", 
                    emp_m.group_name if emp_m else "-", 
                    hired_date_val,
                    d.strftime("%d/%m/%Y"), 
                    ds["first"], ds["last"], ds["std"], display_ot, ds["late"], ds["early"],
                    shift_code_val,
                    ds.get("std_expected", 8.0),
                    ds.get("night_subsidy", 0),
                    # 9 cột phân loại theo shift_category + ca ngày/đêm
                    ds.get("work_normal", 0), ds.get("work_holiday", 0), ds.get("work_rotation", 0),
                    ds.get("ot_normal_day", 0), ds.get("ot_holiday_day", 0), ds.get("ot_rotation_day", 0),
                    ds.get("ot_normal_night", 0), ds.get("ot_holiday_night", 0), ds.get("ot_rotation_night", 0),
                ]
                
                for c_idx, val in enumerate(row_data, 1):
                    cell = ws2.cell(row=ws2_row, column=c_idx, value=val)
                    cell.border = border_style
                    cell.alignment = alignment_style
                
                ws2_row += 1

        with export_lock: export_status["current_step"] = "Generating Sheet 3 (Summary)..."
        ws3_title = f"Bảng công {start_date.strftime('%d.%m')}-{end_date.strftime('%d.%m')}"
        ws3 = wb.create_sheet(title=ws3_title[:31]) # Excel sheet names are limited to 31 chars
        
        ws3_headers = [
            "Mã máy", "Mã công ty", "Tên nhân viên", "Phòng ban", "Nhóm", "Ngày vào làm",
            "Tổng Ngày Công",
            "Nghỉ Phép (P)", "Việc Riêng (R)", "Nghỉ Ốm (O)", "Nghỉ Tang (T)", "Nghỉ Cưới (C)", "Không Phép (K)",
            "Tổng số lần đi muộn", "Tổng số lần về sớm",
            "Tổng Giờ Công", "Tổng Tăng Ca", "Tổng Giờ trợ cấp ca đêm",
            "Tổng Giờ ngày thường", "Tổng Giờ ngày nghỉ lễ", "Tổng Giờ ngày nghỉ luân phiên",
            "Tổng TC thường - ca ngày", "Tổng TC lễ - ca ngày", "Tổng TC luân phiên - ca ngày",
            "Tổng TC thường - ca đêm", "Tổng TC lễ - ca đêm", "Tổng TC luân phiên - ca đêm"
        ]
        ws3.append(ws3_headers)
        for cell in ws3[1]:
            cell.font = bold_font
            cell.alignment = alignment_style
            cell.border = border_style
            cell.fill = header_fill

        for emp_id in sorted_emp_ids:
            emp_info = processed_data[emp_id]
            emp_m = emp_meta.get(emp_id)
            
            late_count = 0
            early_count = 0
            total_workdays = 0.0
            total_hp = total_hr = total_ho = total_ht = total_hc = total_hk = 0.0
            sum_std = sum_ot = sum_ns = 0.0
            sum_wn = sum_wh = sum_wr = 0.0
            sum_otnd = sum_othd = sum_otrd = 0.0
            sum_otnn = sum_othn = sum_otrn = 0.0

            for d in dates_list:
                if d not in emp_info["days"]: continue
                ds = emp_info["days"][d]
                
                # Calculate daily contributions: Actual Hours / Workday Base
                std_h = float(ds.get("std", 0.0))
                base = float(ds.get("workday_base", 8.0))
                if base > 0:
                    total_workdays += (std_h / base)
                    total_hp += (float(ds.get("hp", 0.0)) / base)
                    total_hr += (float(ds.get("hr", 0.0)) / base)
                    total_ho += (float(ds.get("ho", 0.0)) / base)
                    total_ht += (float(ds.get("ht", 0.0)) / base)
                    total_hc += (float(ds.get("hc", 0.0)) / base)
                    total_hk += (float(ds.get("hk", 0.0)) / base)
                
                if isinstance(ds.get("late"), (int, float)) and ds["late"] > 0: late_count += 1
                if isinstance(ds.get("early"), (int, float)) and ds["early"] > 0: early_count += 1
                
                sum_std += float(ds.get("std", 0.0))
                sum_ot  += float(ds.get("ot", 0.0))
                sum_ns  += float(ds.get("night_subsidy", 0.0))
                sum_wn  += float(ds.get("work_normal", 0.0))
                sum_wh  += float(ds.get("work_holiday", 0.0))
                sum_wr  += float(ds.get("work_rotation", 0.0))
                sum_otnd += float(ds.get("ot_normal_day", 0.0))
                sum_othd += float(ds.get("ot_holiday_day", 0.0))
                sum_otrd += float(ds.get("ot_rotation_day", 0.0))
                sum_otnn += float(ds.get("ot_normal_night", 0.0))
                sum_othn += float(ds.get("ot_holiday_night", 0.0))
                sum_otrn += float(ds.get("ot_rotation_night", 0.0))

            full_id_val = emp_info.get("full_id") or "-"
            hired_date_val = emp_m.start_date.strftime("%d/%m/%Y") if emp_m and emp_m.start_date else "-"
            
            row_data_ws3 = [
                emp_id, full_id_val, emp_m.emp_name if emp_m else emp_id, 
                emp_m.department if emp_m else "-", 
                emp_m.group_name if emp_m else "-", 
                hired_date_val,
                round(total_workdays, 3),
                round(total_hp, 3), round(total_hr, 3), round(total_ho, 3),
                round(total_ht, 3), round(total_hc, 3), round(total_hk, 3),
                late_count, early_count,
                round(sum_std, 2), round(sum_ot, 2), round(sum_ns, 2),
                round(sum_wn, 2), round(sum_wh, 2), round(sum_wr, 2),
                round(sum_otnd, 2), round(sum_othd, 2), round(sum_otrd, 2),
                round(sum_otnn, 2), round(sum_othn, 2), round(sum_otrn, 2)
            ]
            ws3.append(row_data_ws3)
            for cell in ws3[ws3.max_row]:
                cell.border = border_style
                cell.alignment = alignment_style

        for sheet in [ws, ws2, ws3]:
            for col_idx in range(1, sheet.max_column + 1):
                column_letter = get_column_letter(col_idx)
                max_length = 0
                for row_idx in range(1, min(sheet.max_row, 20) + 1):
                    val = sheet.cell(row=row_idx, column=col_idx).value
                    if val:
                        length = max(len(str(line)) for line in str(val).split('\n'))
                        if length > max_length: max_length = length
                
                adjusted_width = (max_length + 4) * 1.1 
                sheet.column_dimensions[column_letter].width = min(adjusted_width, 60)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", prefix="Attendance_Export_Temp_")
        wb.save(tmp.name)
        with export_lock:
            export_status.update({"is_running": False, "progress": 100, "filename": tmp.name, "current_step": "Ready for download"})
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        with export_lock:
            export_status.update({"is_running": False, "error": str(e)})
    finally:
        db.close()
